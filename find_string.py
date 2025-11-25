from openpyxl import load_workbook

wb = load_workbook('futuraDesprotegidaModelo1.xlsx', data_only=True)

target = "Interce"
print(f"Searching for '{target}' in all sheets...")

for ws in wb.worksheets:
    print(f"Scanning {ws.title}...")
    for row in ws.iter_rows(max_row=100, max_col=30):
        for cell in row:
            if cell.value and target.lower() in str(cell.value).lower():
                print(f"Found '{cell.value}' in {ws.title} at {cell.coordinate}")
                # Print surrounding context
                r = cell.row
                c = cell.column
                print(f"  Context around {cell.coordinate}:")
                for i in range(max(1, r-2), r+10):
                    val = ws.cell(row=i, column=c).value
                    print(f"    {ws.cell(row=i, column=c).coordinate}: {val}")

