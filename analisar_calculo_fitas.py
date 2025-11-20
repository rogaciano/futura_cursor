"""
Script para analisar aba CÁLCULO e Plan2 (Fitas)
"""
import openpyxl

# Carregar planilha (com fórmulas)
wb = openpyxl.load_workbook('futuraDesprotegidaModelo1.xlsx', data_only=False)
wb_data = openpyxl.load_workbook('futuraDesprotegidaModelo1.xlsx', data_only=True)

print("=" * 80)
print("ANÁLISE DE CÁLCULO E FITAS")
print("=" * 80)

# 1. Analisar Plan2 - Tabela FITAS
print("\n1. ABA 'Plan2' - TABELA FITAS:")
print("-" * 80)
ws_plan2 = wb['Plan2']
ws_plan2_data = wb_data['Plan2']

# O usuário mencionou N33 e intervalo A24:B38
# Vamos ver o que tem nesse intervalo A24:B38 na Plan2 (possivelmente a tabela FITAS)
print("Conteúdo de A24:B40 (Provável tabela FITAS):")
print("Linha | Col A (Largura?) | Col B (Valor?)")
print("-" * 45)
for row in range(24, 41):
    a_val = ws_plan2_data[f'A{row}'].value
    b_val = ws_plan2_data[f'B{row}'].value
    if a_val is not None:
        print(f"  {row}  | {a_val:<16} | {b_val}")
    else:
        print(f"  {row}  | (vazio)          | {b_val}")

# Verificar célula N33 na Plan2
cell_n33 = ws_plan2['N33']
print(f"\nCélula N33 (Plan2) - Fórmula: {cell_n33.value}")
print(f"Célula N33 (Plan2) - Valor:   {ws_plan2_data['N33'].value}")

# 2. Analisar CÁLCULO
print("\n2. ABA 'CÁLCULO':")
print("-" * 80)
ws_calc = wb['CÁLCULO']
ws_calc_data = wb_data['CÁLCULO']

# Verificar B5 (usada na fórmula do PROCV)
b5_formula = ws_calc['B5'].value
b5_value = ws_calc_data['B5'].value
print(f"Célula B5 (CÁLCULO):")
print(f"  Fórmula: {b5_formula}")
print(f"  Valor:   {b5_value}")
print(f"  Descrição (A5?): {ws_calc_data['A5'].value}")

# Olhar ao redor de B5 para contexto
print("\nContexto CÁLCULO (A1:F10):")
for row in range(1, 11):
    vals = []
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        val = ws_calc_data[f'{col}{row}'].value
        vals.append(str(val) if val is not None else "")
    print(f"  L{row}: {' | '.join(vals)}")

# Analisar o "Conversor Unidade x Metros" mencionado na imagem (parece estar no topo à direita)
print("\nProcurando 'CONVERSOR' na aba CÁLCULO:")
found_conversor = False
for row in range(1, 20):
    for col in range(1, 15): # A até N
        cell = ws_calc_data.cell(row=row, column=col)
        if cell.value and 'CONVERSOR' in str(cell.value):
            print(f"  Encontrado em {cell.coordinate}: {cell.value}")
            # Mostrar região ao redor
            for r_offset in range(0, 5):
                r = row + r_offset
                row_vals = []
                for c_offset in range(0, 4):
                    c = col + c_offset
                    val = ws_calc_data.cell(row=r, column=c).value
                    row_vals.append(str(val))
                print(f"    L{r}: {' | '.join(row_vals)}")
            found_conversor = True

if not found_conversor:
    print("  Não encontrado texto 'CONVERSOR' nas primeiras linhas.")

# 3. Verificar se essa tabela de FITAS já existe no banco
print("\n3. VERIFICAÇÃO NO SISTEMA:")
print("-" * 80)
print("Precisamos verificar se esses valores (78.2, 67, 61...) correspondem a algo já cadastrado.")
print("Valores parecem decrescer com a largura.")

wb.close()

