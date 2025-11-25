from openpyxl import load_workbook

try:
    wb = load_workbook('futuraDesprotegidaModelo1.xlsx', data_only=True)
    # List sheet names to be sure
    print(f"Sheets: {wb.sheetnames}")
    
    # Assuming Plan1 as usually that's where main calc is, but user mentioned O23/O25
    # Let's check 'Calculo' or similar if Plan1 is empty/wrong
    ws = wb.active 
    if 'Plan1' in wb.sheetnames:
        ws = wb['Plan1']
        
    print(f"Active sheet: {ws.title}")
    
    val_o23 = ws['O23'].value
    print(f"O23 (Texturas Header?): {val_o23}")
    
    # Texturas items
    print("Texturas values (O24 down):")
    for r in range(24, 40):
        val = ws.cell(row=r, column=15).value # Column 15 is O
        if val:
            print(f"  Row {r}: {val}")
            
    val_o25_header_maybe = ws['O25'].value
    # User said O25 is Materiais. Let's look around O23-O40
    
    # Maybe the table is somewhere else or I need to search for the list the user gave me.
    # "Goma F, Goma G, cola Fria, termocolante..."
    
    # Let's search for "Goma F" in column O
    found = False
    for r in range(1, 100):
        val = ws.cell(row=r, column=15).value
        if str(val).strip() == "Goma F":
            print(f"Found 'Goma F' at O{r}")
            found = True
            # print subsequent rows
            for i in range(1, 10):
                print(f"  + {i}: {ws.cell(row=r+i, column=15).value}")
            break
            
    if not found:
        print("Could not find 'Goma F' in column O. Searching whole sheet...")
        for row in ws.iter_rows(max_row=100, max_col=30):
            for cell in row:
                if cell.value and str(cell.value).strip() == "Goma F":
                     print(f"Found 'Goma F' at {cell.coordinate}")

except Exception as e:
    print(f"Error: {e}")

