from openpyxl import load_workbook

wb = load_workbook('futuraDesprotegidaModelo1.xlsx', data_only=True)
ws = wb['Plan2']

print("Content of Plan2 around E24 (Rows 20-30, Cols A-J):")
for r in range(20, 31):
    row_vals = []
    for c in range(1, 11):
        row_vals.append(str(ws.cell(row=r, column=c).value))
    print(f"Row {r}: {', '.join(row_vals)}")

