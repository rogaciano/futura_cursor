from django.core.management.base import BaseCommand
from django.conf import settings
from decimal import Decimal
from pathlib import Path
from openpyxl import load_workbook
from orcamento.models import Acabamento, PrecoAcabamento

class Command(BaseCommand):
    help = 'Popula a tabela de Acabamentos (Materiais) e Preços baseados na Planilha'

    def handle(self, *args, **options):
        self.stdout.write(self.style.SUCCESS('Iniciando migração de Acabamentos...'))
        
        # 1. Criar Acabamentos
        acabamentos_data = [
            ('Goma F', 'GOMA_F', 1),
            ('Goma G', 'GOMA_G', 2),
            ('Cola Fria', 'COLA_FRIA', 3),
            ('Termocolante', 'TERMOCOLANTE', 4),
            ('Interce Fino', 'INTERCE_FINO', 5),
            ('Interce Grosso', 'INTERCE_GROSSO', 6),
            ('Overloque + interce', 'OVERLOQUE_INTERCE', 7),
        ]
        
        acabamento_objs = {}
        
        for nome, codigo, ordem in acabamentos_data:
            obj, created = Acabamento.objects.get_or_create(
                codigo=codigo,
                defaults={'nome': nome, 'ordem': ordem}
            )
            acabamento_objs[codigo] = obj
            if created:
                self.stdout.write(f'Criado Acabamento: {nome}')
            else:
                self.stdout.write(f'Acabamento existente: {nome}')
        
        # 2. Carregar Preços da Planilha (Plan2)
        base_dir = Path(settings.BASE_DIR)
        caminho = base_dir / 'futuraDesprotegidaModelo1.xlsx'
        
        if not caminho.exists():
            self.stdout.write(self.style.ERROR('Planilha não encontrada!'))
            return

        wb = load_workbook(caminho, data_only=True)
        if 'Plan2' not in wb.sheetnames:
            self.stdout.write(self.style.ERROR('Plan2 não encontrada!'))
            return
            
        ws = wb['Plan2']
        
        # Mapeamento de Colunas (conforme verificado via script inspect_plan2.py)
        # Row 23 headers: LARGURA, F, G, colafria, termo, fino, grosso
        # Col indices (1-based):
        # D (4): Largura
        # E (5): F -> GOMA_F
        # F (6): G -> GOMA_G
        # G (7): colafria -> COLA_FRIA
        # H (8): termo -> TERMOCOLANTE
        # I (9): fino -> INTERCE_FINO
        # J (10): grosso -> INTERCE_GROSSO
        
        col_map = {
            5: 'GOMA_F',
            6: 'GOMA_G',
            7: 'COLA_FRIA',
            8: 'TERMOCOLANTE',
            9: 'INTERCE_FINO',
            10: 'INTERCE_GROSSO',
        }
        
        count_created = 0
        
        # Iterar linhas a partir da 24 (onde começam os dados)
        for r in range(24, 100):
            largura = ws.cell(row=r, column=4).value
            if not largura:
                break # Fim da tabela
                
            try:
                largura = int(largura)
            except:
                continue
                
            for col_idx, cod_acabamento in col_map.items():
                valor = ws.cell(row=r, column=col_idx).value
                if valor is None:
                    valor = 0
                    
                try:
                    preco = Decimal(str(valor))
                except:
                    preco = Decimal('0.0')
                
                acabamento = acabamento_objs[cod_acabamento]
                
                PrecoAcabamento.objects.update_or_create(
                    largura_mm=largura,
                    acabamento=acabamento,
                    defaults={'preco': preco}
                )
                count_created += 1
        
        self.stdout.write(self.style.SUCCESS(f'Sucesso! {count_created} preços importados.'))
        
        # Nota sobre Overloque
        self.stdout.write(self.style.WARNING('ATENÇÃO: Preços para "Overloque + interce" não foram importados (não encontrados na Plan2).'))

