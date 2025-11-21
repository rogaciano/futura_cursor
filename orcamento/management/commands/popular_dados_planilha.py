from django.core.management.base import BaseCommand
from django.conf import settings
from decimal import Decimal
from pathlib import Path
from openpyxl import load_workbook
from orcamento.models import (
    TipoMaterial, TipoCorte, TabelaPreco, CoeficienteFator,
    ValorGoma, ValorCorte, Configuracao, Textura
)


class Command(BaseCommand):
    help = 'Popula o banco de dados com os dados da planilha Excel'

    def handle(self, *args, **options):
        self.stdout.write(self.style.SUCCESS('Iniciando população do banco de dados...'))
        base_dir = Path(settings.BASE_DIR)
        
        # 1. Criar Tipos de Material
        self.stdout.write('Criando tipos de material...')
        # Formato: (nome, codigo, ordem)
        materiais = [
            ('Tafetá', 'TAFETA', 1),
            ('Sarja', 'SARJA', 2),
            ('Alta Definição', 'ALTA_DEF', 3),
            ('Dupla Densidade', 'DUPLA_DENS', 4),
            ('Super Batidas', 'SUPER_BAT', 5),
            ('Canvas', 'CANVAS', 6),
            ('Cetim', 'CETIM', 7),
            ('SuperSoft', 'SUPERSOFT', 8),
        ]
        
        for nome, codigo, ordem in materiais:
            material, created = TipoMaterial.objects.update_or_create(
                codigo=codigo,
                defaults={
                    'nome': nome,
                    'ordem': ordem,
                    'ativo': True,
                }
            )
            if created:
                self.stdout.write(f'  [+] Criado: {nome}')
            else:
                self.stdout.write(f'  [~] Atualizado: {nome}')
        
        # 2. Criar Tipos de Corte
        self.stdout.write('Criando tipos de corte...')
        cortes = [
            ('ETQ SCORTE', 'ETQ_SCORTE', 5),
            ('CORTE NORMAL', 'CORTE_NORMAL', 6),
            ('DOBRA MEIO', 'DOBRA_MEIO', 7),
            ('DOBRA CANTOS', 'DOBRA_CANTOS', 8),
            ('CANTOSDIAGONAL', 'CANTOS_DIAGONAL', 9),
            ('DOBRA DESCENTRALIZADA', 'DOBRA_DESC', 10),
            ('ENVELOPE', 'ENVELOPE', 11),
            ('MITRA', 'MITRA', 12),
            ('2 DOBRAS', 'DOIS_DOBRAS', 13),
            # cortes auxiliares já existentes para compatibilidade
            ('MEIO CORTE', 'MEIO_CORTE', 14),
            ('SAQUINHO', 'SAQUINHO', 15),
            ('CORTE ESPECIAL', 'CORTE_ESP', 16),
        ]
        
        for nome, codigo, codigo_calc in cortes:
            TipoCorte.objects.get_or_create(
                codigo=codigo,
                defaults={'nome': nome, 'codigo_calc': codigo_calc}
            )
        
        # 3. Criar Tabela de Preços (baseado na planilha Plan2)
        self.stdout.write('Criando tabela de preços...')
        metragens = [300, 500, 1000, 2500, 5000, 10000, 15000, 30000]
        
        # Preços exemplo para Tafetá
        precos_tafeta = [20.40, 15.69, 10.44, 10.11, 9.35, 8.39, 8.04, 7.72]
        tafeta = TipoMaterial.objects.get(codigo='TAFETA')
        for metragem, preco in zip(metragens, precos_tafeta):
            TabelaPreco.objects.get_or_create(
                metragem=metragem,
                tipo_material=tafeta,
                defaults={'preco_metro': Decimal(str(preco))}
            )
        
        # Preços exemplo para Sarja
        precos_sarja = [22.50, 17.29, 11.48, 11.12, 10.29, 9.22, 8.84, 8.48]
        sarja = TipoMaterial.objects.get(codigo='SARJA')
        for metragem, preco in zip(metragens, precos_sarja):
            TabelaPreco.objects.get_or_create(
                metragem=metragem,
                tipo_material=sarja,
                defaults={'preco_metro': Decimal(str(preco))}
            )
        
        # Preços exemplo para Alta Definição
        precos_alta_def = [26.84, 20.64, 14.36, 13.21, 12.58, 11.89, 11.35, 10.89]
        alta_def = TipoMaterial.objects.get(codigo='ALTA_DEF')
        for metragem, preco in zip(metragens, precos_alta_def):
            TabelaPreco.objects.get_or_create(
                metragem=metragem,
                tipo_material=alta_def,
                defaults={'preco_metro': Decimal(str(preco))}
            )
        
        # 4. Criar Coeficientes Fator (baseados na planilha)
        self.stdout.write('Criando coeficientes fator...')
        coef_planilha = self._carregar_coeficientes_planilha(base_dir)
        larguras_disponiveis = sorted(coef_planilha.pop('_larguras', set()))
        if not larguras_disponiveis:
            larguras_disponiveis = [10, 12, 15, 18, 20, 21, 24, 28, 30, 33, 40, 50, 67, 100]

        materiais_qs = TipoMaterial.objects.all()
        cortes_qs = TipoCorte.objects.all()

        for largura in larguras_disponiveis:
            for corte in cortes_qs:
                for material in materiais_qs:
                    valor = coef_planilha.get(
                        (corte.nome.upper(), material.codigo.upper()),
                        {}
                    ).get(largura, Decimal('1.0'))

                    CoeficienteFator.objects.update_or_create(
                        largura=largura,
                        tipo_material=material,
                        codigo_corte=corte,
                        defaults={'coeficiente': Decimal(str(valor))}
                    )
        
        # 5. Criar Valores de Goma (baseado na planilha)
        self.stdout.write('Criando valores de goma...')
        larguras_goma = [10, 12, 15, 18, 20, 21, 24, 28, 30, 33, 40, 50, 67, 100]
        gomas_fino = [0.029, 0.035, 0.043, 0.051, 0.058, 0.065, 0.073, 0.085, 0.1, 0.1, 0.116, 0.145, 0.194, 0.29]
        gomas_grosso = [0.052, 0.062, 0.078, 0.089, 0.1, 0.107, 0.11, 0.128, 0.136, 0.136, 0.176, 0.22, 0.295, 0.44]
        gomas_termo = [0.2, 0.2, 0.25, 0.3, 0.3, 0.3, 0.4, 0.45, 0.5, 0.5, 0.55, 0.7, 0.95, 1.4]
        
        for largura, fino, grosso, termo in zip(larguras_goma, gomas_fino, gomas_grosso, gomas_termo):
            ValorGoma.objects.get_or_create(
                largura=largura,
                defaults={
                    'goma_fino': Decimal(str(fino)),
                    'goma_grosso': Decimal(str(grosso)),
                    'termocolante': Decimal(str(termo))
                }
            )
        
        # 6. Criar Valores de Corte (Canvas e Cetim)
        self.stdout.write('Criando valores de corte especial...')
        for largura in larguras_goma:
            ValorCorte.objects.get_or_create(
                largura=largura,
                defaults={
                    'canvas': Decimal('0.189'),
                    'cetim': Decimal('0.087')
                }
            )
        
        # 7. Criar Configurações
        self.stdout.write('Criando configurações...')
        configs = [
            ('perc_ultrassonico', '1.15', 'Percentual de aumento para corte ultrassônico', 'decimal'),
            ('perc_aumento_geral', '1.00', 'Percentual de aumento geral', 'decimal'),
        ]
        
        for chave, valor, descricao, tipo_dado in configs:
            Configuracao.objects.get_or_create(
                chave=chave,
                defaults={
                    'valor': valor,
                    'descricao': descricao,
                    'tipo_dado': tipo_dado
                }
            )
        
        # 8. Criar Texturas (baseado na planilha)
        self.stdout.write('Criando texturas...')
        texturas = [
            ('1', 'PIMENTAS'),
            ('2', 'FOLHAS'),
            ('3', 'HIBISCO'),
            ('4', 'CAFÉ'),
            ('5', 'FLORES'),
            ('6', 'PEDRAS'),
            ('7', 'XADREZ'),
            ('8', 'LIG8'),
            ('9', 'LIG10'),
            ('10', 'LIG16'),
            ('11', 'SARJADO'),
            ('12', 'ZIG ZAG'),
            ('13', 'ALMOFADADO'),
            ('14', 'ALMOFADADO P'),
            ('15', 'DIAMANTADO P'),
            ('16', 'LADRILHO'),
            ('17', 'MOSAICO'),
            ('18', 'INDIANO'),
            ('19', 'CANELADO'),
            ('20', 'ONDAS'),
            ('21', 'CASULO'),
            ('22', 'CIRCULOS'),
            ('23', 'DIAMANTADO'),
            ('24', 'DEGRAUS'),
            ('25', 'DUNAS'),
            ('26', 'ENGRENAGEM'),
            ('27', 'JEANS'),
            ('28', 'SAFIRA'),
            ('29', 'CORAÇÃO'),
            ('30', '3D'),
        ]
        
        for ordem, (codigo, nome) in enumerate(texturas, 1):
            Textura.objects.get_or_create(
                codigo=codigo,
                defaults={'nome': nome, 'ordem': ordem}
            )
        
        self.stdout.write(self.style.SUCCESS('Banco de dados populado com sucesso!'))
        self.stdout.write(self.style.SUCCESS(f'  - {TipoMaterial.objects.count()} tipos de material'))
        self.stdout.write(self.style.SUCCESS(f'  - {TipoCorte.objects.count()} tipos de corte'))
        self.stdout.write(self.style.SUCCESS(f'  - {TabelaPreco.objects.count()} precos cadastrados'))
        self.stdout.write(self.style.SUCCESS(f'  - {CoeficienteFator.objects.count()} coeficientes'))
        self.stdout.write(self.style.SUCCESS(f'  - {ValorGoma.objects.count()} valores de goma'))
        self.stdout.write(self.style.SUCCESS(f'  - {ValorCorte.objects.count()} valores de corte'))
        self.stdout.write(self.style.SUCCESS(f'  - {Configuracao.objects.count()} configuracoes'))
        self.stdout.write(self.style.SUCCESS(f'  - {Textura.objects.count()} texturas'))

    def _carregar_coeficientes_planilha(self, base_dir: Path):
        """Extrai o mapa de coeficientes diretamente da aba Plan2 da planilha."""
        caminho = base_dir / 'futuraDesprotegidaModelo1.xlsx'
        dados = {}
        larguras = set()

        if not caminho.exists():
            self.stdout.write(self.style.WARNING('Planilha "futuraDesprotegidaModelo1.xlsx" não encontrada. Coeficientes padrão (1.0) serão usados.'))
            dados['_larguras'] = larguras
            return dados

        wb = load_workbook(caminho, data_only=True)
        if 'Plan2' not in wb.sheetnames:
            self.stdout.write(self.style.WARNING('A planilha não possui aba "Plan2".'))
            dados['_larguras'] = larguras
            return dados

        ws = wb['Plan2']

        grupos = {}
        header_row = 3
        for col in range(9, 40, 3):
            nome = ws.cell(row=header_row, column=col).value
            if nome:
                grupos[col] = str(nome).strip().upper()

        material_codes = ['TAFETA', 'SARJA', 'ALTA_DEF']

        for row in range(5, 40):
            largura = ws.cell(row=row, column=8).value
            if not largura:
                continue

            try:
                largura = int(largura)
            except (TypeError, ValueError):
                continue

            larguras.add(largura)

            for col_base, corte_nome in grupos.items():
                for offset, material_code in enumerate(material_codes):
                    valor = ws.cell(row=row, column=col_base + offset).value
                    if valor is None:
                        continue

                    try:
                        decimal_value = Decimal(str(valor))
                    except Exception:
                        continue

                    dados.setdefault((corte_nome, material_code), {})[largura] = decimal_value

        dados['_larguras'] = larguras
        return dados

