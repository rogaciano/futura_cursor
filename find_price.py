from openpyxl import load_workbook
import sys

wb = load_workbook('futuraDesprotegidaModelo1.xlsx', data_only=True)

# Goma Fino first value is 0.029
target_val = 0.029

print(f"Searching for value {target_val}...")

for ws in wb.worksheets:
    for row in ws.iter_rows(max_row=100, max_col=30):
        for cell in row:
            try:
                # Approximate comparison for floats
                if cell.value and isinstance(cell.value, (int, float)) and abs(cell.value - target_val) < 0.0001:
                    print(f"Found {target_val} in {ws.title} at {cell.coordinate}")
                    # Show context (headers usually above or left)
                    r = cell.row
                    c = cell.column
                    # Print header row (row 3 or 4 usually)
                    print(f"  Header check row 3, col {c}: {ws.cell(row=3, column=c).value}")
                    print(f"  Header check row 4, col {c}: {ws.cell(row=4, column=c).value}")
            except:
                pass

