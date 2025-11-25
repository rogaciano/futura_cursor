from openpyxl import load_workbook

try:
    wb = load_workbook('futuraDesprotegidaModelo1.xlsx', data_only=True)
    
    # Try CLCULO sheet (might need encoding handling or just get by index)
    # It's the 3rd sheet (index 2)
    ws = wb.worksheets[2]
    print(f"Sheet: {ws.title}")
    
    val_o23 = ws['O23'].value
    print(f"O23: {val_o23}")
    
    val_o25 = ws['O25'].value
    print(f"O25: {val_o25}")

    # Look for the list in Column O
    print("Scanning Column O (15)...")
    for r in range(1, 100):
        val = ws.cell(row=r, column=15).value
        if val:
            print(f"O{r}: {val}")

except Exception as e:
    print(f"Error: {e}")

